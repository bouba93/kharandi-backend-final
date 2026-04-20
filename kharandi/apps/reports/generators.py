"""
╔══════════════════════════════════════════════════════════════════════════╗
║           KHARANDI — Rapports : PDF (ReportLab) + Excel (OpenPyXL)     ║
╚══════════════════════════════════════════════════════════════════════════╝
"""

import io
from django.http import FileResponse
from django.utils import timezone
from rest_framework.views import APIView
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response

from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable
from reportlab.lib.enums import TA_CENTER

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ── Couleurs ──────────────────────────────────────────────────────────────
BLUE  = colors.HexColor("#1a73e8")
LBLUE = colors.HexColor("#e8f0fe")
GRAY  = colors.HexColor("#666666")


def _header(elements, title: str, subtitle: str = ""):
    s1 = ParagraphStyle("T", fontSize=22, fontName="Helvetica-Bold",
                         textColor=BLUE, alignment=TA_CENTER, spaceAfter=4)
    s2 = ParagraphStyle("S", fontSize=11, fontName="Helvetica",
                         textColor=GRAY, alignment=TA_CENTER, spaceAfter=2)
    s3 = ParagraphStyle("D", fontSize=9,  fontName="Helvetica",
                         textColor=colors.HexColor("#999"), alignment=TA_CENTER)
    s4 = ParagraphStyle("H", fontSize=16, fontName="Helvetica-Bold",
                         textColor=colors.HexColor("#222"), alignment=TA_CENTER, spaceAfter=12)
    elements.append(Paragraph("KHARANDI", s1))
    if subtitle: elements.append(Paragraph(subtitle, s2))
    elements.append(Paragraph(f"Généré le {timezone.now().strftime('%d/%m/%Y à %H:%M')} (Conakry)", s3))
    elements.append(Spacer(1, 0.3*cm))
    elements.append(HRFlowable(width="100%", thickness=2, color=BLUE))
    elements.append(Spacer(1, 0.5*cm))
    elements.append(Paragraph(title, s4))


def _table_style():
    return TableStyle([
        ("BACKGROUND",    (0, 0),  (-1, 0),  BLUE),
        ("TEXTCOLOR",     (0, 0),  (-1, 0),  colors.white),
        ("FONTNAME",      (0, 0),  (-1, 0),  "Helvetica-Bold"),
        ("FONTSIZE",      (0, 0),  (-1, 0),  10),
        ("ALIGN",         (0, 0),  (-1, -1), "CENTER"),
        ("VALIGN",        (0, 0),  (-1, -1), "MIDDLE"),
        ("FONTNAME",      (0, -1), (-1, -1), "Helvetica-Bold"),
        ("BACKGROUND",    (0, -1), (-1, -1), LBLUE),
        ("GRID",          (0, 0),  (-1, -1), 0.5, colors.HexColor("#cccccc")),
        ("FONTSIZE",      (0, 1),  (-1, -1), 9),
        ("TOPPADDING",    (0, 0),  (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0),  (-1, -1), 6),
        ("ROWBACKGROUNDS",(0, 1),  (-1, -2), [colors.white, LBLUE]),
    ])


def _make_doc(buf):
    return SimpleDocTemplate(buf, pagesize=A4,
                             leftMargin=2*cm, rightMargin=2*cm,
                             topMargin=2*cm,  bottomMargin=2*cm)


# ══════════════════════════════════════════════════════════════════════════
#  PDF TRANSACTIONS VENDEUR
# ══════════════════════════════════════════════════════════════════════════

def generate_transactions_pdf_buffer(transactions: list, vendor_name: str) -> io.BytesIO:
    buf = io.BytesIO()
    doc = _make_doc(buf)
    el  = []
    _header(el, f"Relevé de transactions — {vendor_name}", "Historique des ventes sur Kharandi")

    if not transactions:
        el.append(Paragraph("Aucune transaction.", getSampleStyleSheet()["Normal"]))
    else:
        total = sum(t.get("amount", 0) for t in transactions if t.get("status") == "SUCCESS")
        rows  = [["Date", "Commande", "Produit", "Qté", "Montant (GNF)", "Commission", "Net (GNF)", "Statut"]]
        for t in transactions:
            rows.append([
                t.get("date", "-"),
                f"#{t.get('order_id', '-')}",
                t.get("product", "-")[:30],
                str(t.get("qty", 1)),
                f"{t.get('amount', 0):,}",
                f"{t.get('commission', 0):,}",
                f"{t.get('net', 0):,}",
                "Succès" if t.get("status") == "SUCCESS" else "Échec",
            ])
        rows.append(["", "", "", "TOTAL :", f"{total:,}", "", "", ""])
        tbl = Table(rows, colWidths=[2.3*cm, 2.3*cm, 3.5*cm, 1*cm, 2.5*cm, 2.2*cm, 2.2*cm, 1.5*cm], repeatRows=1)
        tbl.setStyle(_table_style())
        el.append(tbl)
    doc.build(el)
    buf.seek(0)
    return buf


# ══════════════════════════════════════════════════════════════════════════
#  PDF BULLETIN DE NOTES
# ══════════════════════════════════════════════════════════════════════════

def generate_student_report_pdf_buffer(student: dict, courses: list) -> io.BytesIO:
    buf = io.BytesIO()
    doc = _make_doc(buf)
    el  = []
    _header(el, "Bulletin de Notes", "Rapport de performance scolaire — Kharandi")

    info = [
        ["Élève :", student.get("name", "-"),   "Classe :", student.get("level", "-")],
        ["École :", student.get("school", "-"),  "Année :",  str(timezone.now().year)],
    ]
    itbl = Table(info, colWidths=[3*cm, 6*cm, 3*cm, 5*cm])
    itbl.setStyle(TableStyle([
        ("FONTNAME",      (0,0),(0,-1), "Helvetica-Bold"),
        ("FONTNAME",      (2,0),(2,-1), "Helvetica-Bold"),
        ("TEXTCOLOR",     (0,0),(0,-1), BLUE),
        ("TEXTCOLOR",     (2,0),(2,-1), BLUE),
        ("FONTSIZE",      (0,0),(-1,-1), 10),
        ("BOTTOMPADDING", (0,0),(-1,-1), 5),
    ]))
    el.append(itbl)
    el.append(Spacer(1, 0.5*cm))
    el.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor("#ddd")))
    el.append(Spacer(1, 0.4*cm))

    rows  = [["Matière", "Enseignant", "Note", "/ Max", "Mention"]]
    total = 0
    for c in courses:
        s  = c.get("score", 0)
        mx = c.get("max_score", 20)
        total += (float(s) / float(mx)) * 20 if mx else 0
        rows.append([c.get("subject","-"), c.get("teacher","-"), str(s), str(mx), c.get("grade","-")])
    moy = total / len(courses) if courses else 0
    rows.append(["", "", f"MOY: {moy:.2f}", "/20", ""])

    tbl = Table(rows, colWidths=[4*cm, 4.5*cm, 2*cm, 2*cm, 4.5*cm], repeatRows=1)
    tbl.setStyle(_table_style())
    el.append(tbl)
    doc.build(el)
    buf.seek(0)
    return buf


# ══════════════════════════════════════════════════════════════════════════
#  PDF FACTURE
# ══════════════════════════════════════════════════════════════════════════

def generate_invoice_pdf_buffer(transaction) -> io.BytesIO:
    buf = io.BytesIO()
    doc = _make_doc(buf)
    el  = []
    _header(el, "FACTURE", "Kharandi — Reçu de paiement")

    rows = [
        ["Numéro LengoPay :", transaction.lengopay_id or "-"],
        ["Payeur :",          transaction.payer.get_full_name()],
        ["Téléphone :",       str(transaction.payer.phone)],
        ["Date :",            transaction.created_at.strftime("%d/%m/%Y %H:%M")],
        ["Montant :",         f"{int(transaction.amount):,} GNF"],
        ["Commission Kharandi :", f"{int(transaction.commission_amount):,} GNF"],
        ["Montant net vendeur :", f"{int(transaction.net_amount):,} GNF"],
        ["Statut :",          "✅ PAYÉ" if transaction.status == "success" else transaction.status.upper()],
    ]
    tbl = Table(rows, colWidths=[7*cm, 10*cm])
    tbl.setStyle(TableStyle([
        ("FONTNAME",      (0,0),(0,-1), "Helvetica-Bold"),
        ("TEXTCOLOR",     (0,0),(0,-1), BLUE),
        ("FONTSIZE",      (0,0),(-1,-1), 10),
        ("BOTTOMPADDING", (0,0),(-1,-1), 8),
        ("ROWBACKGROUNDS",(0,0),(-1,-1), [colors.white, LBLUE]),
        ("GRID",          (0,0),(-1,-1), 0.5, colors.HexColor("#eeeeee")),
    ]))
    el.append(tbl)
    doc.build(el)
    buf.seek(0)
    return buf


# ══════════════════════════════════════════════════════════════════════════
#  EXCEL STATS
# ══════════════════════════════════════════════════════════════════════════

def generate_stats_excel_buffer(stats: dict) -> io.BytesIO:
    buf = io.BytesIO()
    wb  = openpyxl.Workbook()

    HX_BLUE  = "1a73e8"
    HX_LBLUE = "e8f0fe"
    HX_WHITE = "FFFFFF"
    HX_DARK  = "222222"
    thin = Side(style="thin", color="CCCCCC")
    brd  = Border(left=thin, right=thin, top=thin, bottom=thin)

    def h(cell, bg=HX_BLUE, fg=HX_WHITE):
        cell.font      = Font(bold=True, color=fg, size=11)
        cell.fill      = PatternFill("solid", fgColor=bg)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = brd

    def r(cell, alt=False):
        cell.fill      = PatternFill("solid", fgColor=HX_LBLUE if alt else HX_WHITE)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.font      = Font(color=HX_DARK, size=10)
        cell.border    = brd

    def sheet(wb, title, cols, rows_data, is_first=False):
        ws = wb.active if is_first else wb.create_sheet(title)
        ws.title = title
        ws.row_dimensions[1].height = 30
        for ci, c in enumerate(cols, 1):
            h(ws.cell(row=1, column=ci, value=c))
            ws.column_dimensions[chr(64+ci) if ci <= 26 else "A"].width = 22
        for ri, row in enumerate(rows_data, 2):
            for ci, v in enumerate(row, 1):
                r(ws.cell(row=ri, column=ci, value=v), alt=(ri % 2 == 0))
        return ws

    # Utilisateurs
    users = stats.get("users", [])
    sheet(wb, "Utilisateurs",
          ["ID", "Nom", "Rôle", "Points", "Téléphone", "Ville", "Statut", "Inscription"],
          [[u.get("id",""), u.get("name",""), u.get("role",""), u.get("points",0),
            u.get("phone",""), u.get("city",""), u.get("status",""), u.get("created_at","")] for u in users],
          is_first=True)

    # Ventes
    sales     = stats.get("sales", [])
    ws2       = sheet(wb, "Ventes",
                      ["Order ID","Produit","Vendeur","Acheteur","Montant","Commission","Net","Statut","Date"],
                      [[s.get("order_id",""), s.get("product",""), s.get("seller",""), s.get("buyer",""),
                        s.get("amount",0), s.get("commission",0), s.get("net",0), s.get("status",""), s.get("date","")] for s in sales])
    total_ventes = sum(s.get("amount",0) for s in sales if s.get("status") == "SUCCESS")
    last = len(sales) + 2
    ws2.cell(row=last, column=4, value="TOTAL :").font = Font(bold=True, color=HX_BLUE)
    c = ws2.cell(row=last, column=5, value=total_ventes)
    c.font = Font(bold=True, color=HX_BLUE)
    c.fill = PatternFill("solid", fgColor=HX_LBLUE)

    # Cours
    sheet(wb, "Cours",
          ["ID","Titre","Matière","Répétiteur","Niveau","Élèves","Note","Prix"],
          [[c.get("id",""), c.get("title",""), c.get("subject",""), c.get("teacher",""),
            c.get("level",""), c.get("students_count",0), c.get("avg_rating",0), c.get("price",0)]
           for c in stats.get("courses", [])])

    # Commissions
    sheet(wb, "Commissions",
          ["Transaction","Vendeur","Taux %","Brut","Commission","Net","Payé","Date"],
          [[c.get("txn_id",""), c.get("vendor",""), f"{float(c.get('rate',0))*100:.1f}%",
            c.get("gross",0), c.get("commission",0), c.get("net",0),
            "Oui" if c.get("paid") else "Non", c.get("date","")]
           for c in stats.get("commissions", [])])

    wb.save(buf)
    buf.seek(0)
    return buf


# ══════════════════════════════════════════════════════════════════════════
#  VUES DRF
# ══════════════════════════════════════════════════════════════════════════

class TransactionsPDFView(APIView):
    permission_classes = [IsAuthenticated]
    def post(self, request):
        name = request.data.get("vendor_name", request.user.get_full_name())
        txns = request.data.get("transactions", [])
        buf  = generate_transactions_pdf_buffer(txns, name)
        fn   = f"kharandi_transactions_{timezone.now().strftime('%Y%m%d')}.pdf"
        return FileResponse(buf, as_attachment=True, filename=fn, content_type="application/pdf")


class StudentBulletinPDFView(APIView):
    permission_classes = [IsAuthenticated]
    def post(self, request):
        student = request.data.get("student", {})
        courses = request.data.get("courses", [])
        if not student.get("name"):
            return Response({"success": False, "message": "Champ 'student.name' requis."}, status=400)
        buf = generate_student_report_pdf_buffer(student, courses)
        fn  = f"bulletin_{student['name'].replace(' ','_')}_{timezone.now().strftime('%Y%m')}.pdf"
        return FileResponse(buf, as_attachment=True, filename=fn, content_type="application/pdf")


class StatsExcelView(APIView):
    permission_classes = [IsAuthenticated]
    def post(self, request):
        if not (request.user.is_staff or request.user.role == "admin"):
            return Response({"success": False, "message": "Accès réservé aux admins."}, status=403)
        buf = generate_stats_excel_buffer(request.data)
        fn  = f"kharandi_stats_{timezone.now().strftime('%Y%m%d')}.xlsx"
        return FileResponse(buf, as_attachment=True, filename=fn,
                            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


class MyTransactionsPDFView(APIView):
    """Génère le relevé PDF pour l'utilisateur connecté depuis ses vraies transactions."""
    permission_classes = [IsAuthenticated]
    def get(self, request):
        from kharandi.apps.payments.models import Transaction
        txns = Transaction.objects.filter(
            payer=request.user, status="success"
        ).select_related("order").order_by("-created_at")[:100]

        data = []
        for t in txns:
            data.append({
                "date":       t.created_at.strftime("%d/%m/%Y"),
                "order_id":   t.order.order_number if t.order else "-",
                "product":    f"Commande #{t.order.order_number}" if t.order else t.transaction_type,
                "qty":        1,
                "amount":     int(t.amount),
                "commission": int(t.commission_amount),
                "net":        int(t.net_amount),
                "status":     "SUCCESS",
            })

        buf = generate_transactions_pdf_buffer(data, request.user.get_full_name())
        fn  = f"mes_transactions_{timezone.now().strftime('%Y%m%d')}.pdf"
        return FileResponse(buf, as_attachment=True, filename=fn, content_type="application/pdf")
